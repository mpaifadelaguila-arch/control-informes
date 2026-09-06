import io
import json
import os
import re
import threading
import time
from datetime import datetime

import pandas as pd
import streamlit as st
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload, MediaIoBaseDownload
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Configuración de página
st.set_page_config(
    page_title="Control interno de informes - Ademinsac",
    page_icon=":material/assignment:",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# Prevención del error removeChild (bloquea la traducción automática del navegador que corrompe el DOM de React)
st.markdown(
    """
    <head>
        <meta name="google" content="notranslate">
    </head>
    """,
    unsafe_allow_html=True
)

# Constantes de Google Drive
FOLDER_ID = "1gUyx6PbtLd7tG_C20x00CVmVdF0oYm_8"

@st.cache_resource
def conectar_drive():
    try:
        scopes = ['https://www.googleapis.com/auth/drive']
        creds_dict = dict(st.secrets["gcp_service_account"])
        creds_dict["private_key"] = creds_dict["private_key"].replace("\\n", "\n").replace("\r\n", "\n")
        credentials = Credentials.from_service_account_info(creds_dict, scopes=scopes)
        service = build('drive', 'v3', credentials=credentials)
        return service
    except Exception as e:
        st.error(f"Error al conectar con Google Drive: {e}")
        return None

drive_service = conectar_drive()

def descargar_archivo_de_drive(nombre_archivo, ruta_local, max_reintentos=3):
    """Descarga la versión más reciente desde Drive con reintentos automáticos para evitar errores SSL/Network."""
    if not drive_service:
        return False
        
    for intento in range(1, max_reintentos + 1):
        try:
            query = f"'{FOLDER_ID}' in parents and name = '{nombre_archivo}' and trashed = false"
            res = drive_service.files().list(
                q=query, 
                fields="files(id)",
                supportsAllDrives=True,
                includeItemsFromAllDrives=True
            ).execute()
            archivos = res.get('files', [])

            if archivos:
                file_id = archivos[0]['id']
                request = drive_service.files().get_media(fileId=file_id)
                with open(ruta_local, 'wb') as f:
                    downloader = MediaIoBaseDownload(f, request)
                    done = False
                    while not done:
                        _, done = downloader.next_chunk()
                return True
            return False
        except Exception as e:
            msg_error = str(e)
            if ("RECORD_LAYER_FAILURE" in msg_error or "SSL" in msg_error or "Connection" in msg_error) and intento < max_reintentos:
                time.sleep(1.2 * intento)  # Espera exponencial progresiva
                continue
            st.error(f"Error al descargar desde Google Drive ({nombre_archivo}): {e}")
            break
    return False

def subir_archivo_a_drive(nombre_archivo, ruta_local, mime_type='application/json', max_reintentos=3):
    """Subida síncrona a Google Drive con reintentos automáticos contra fallos de socket SSL."""
    if not drive_service:
        return False
        
    for intento in range(1, max_reintentos + 1):
        try:
            query = f"'{FOLDER_ID}' in parents and name = '{nombre_archivo}' and trashed = false"
            res = drive_service.files().list(
                q=query, 
                fields="files(id)",
                supportsAllDrives=True,
                includeItemsFromAllDrives=True
            ).execute()
            archivos = res.get('files', [])

            with open(ruta_local, 'rb') as f:
                contenido_binario = io.BytesIO(f.read())

            media = MediaIoBaseUpload(contenido_binario, mimetype=mime_type, resumable=True)

            if archivos:
                file_id = archivos[0]['id']
                drive_service.files().update(
                    fileId=file_id, 
                    media_body=media,
                    supportsAllDrives=True
                ).execute()
            else:
                file_metadata = {
                    'name': nombre_archivo, 
                    'parents': [FOLDER_ID]
                }
                drive_service.files().create(
                    body=file_metadata, 
                    media_body=media,
                    supportsAllDrives=True,
                    fields='id'
                ).execute()

            return True
        except Exception as e:
            msg_error = str(e)
            if ("RECORD_LAYER_FAILURE" in msg_error or "SSL" in msg_error or "Connection" in msg_error) and intento < max_reintentos:
                time.sleep(1.2 * intento)
                continue
            print(f"Error al respaldar en Google Drive ({nombre_archivo}): {e}")
            break
    return False

def subir_a_drive_en_segundo_plano(nombre_archivo, ruta_local, mime_type='application/json'):
    """Ejecuta la subida a Drive en un hilo secundario para evitar bloqueos en la interfaz."""
    hilo = threading.Thread(
        target=subir_archivo_a_drive,
        args=(nombre_archivo, ruta_local, mime_type),
        daemon=True
    )
    hilo.start()

# Estilos CSS Corporativos
st.markdown(
    """
    <style>
    footer {visibility: hidden;}
    .stAppDeployButton {display:none !important;}
    header {visibility: hidden !important;}

    :root {
        --primary-navy: #0E2A47;
        --secondary-navy: #1A3E68;
        --gold-accent: #D4AF37;
        --bg-card: #FFFFFF;
        --border-color: #E2E8F0;
        --text-main: #1E293B;
        --text-sub: #64748B;
    }

    .stApp { background-color: #EEF2F7; }
    .block-container { padding-top: 1.6rem !important; }

    /* HEADER BANNER */
    .header-banner {
        background: linear-gradient(120deg, #0B2038 0%, #1E4E7E 60%, #2C6494 100%);
        padding: 22px 30px;
        border-radius: 14px;
        color: white;
        margin-bottom: 20px;
        box-shadow: 0 12px 28px rgba(11, 32, 56, 0.18);
        position: relative;
        overflow: hidden;
    }
    .header-banner::after {
        content: "";
        position: absolute; top: 0; right: 0; bottom: 0; width: 6px;
        background: linear-gradient(180deg, #E7BE30, #C99A1E);
    }
    .header-title { font-size: 24px; font-weight: 800; letter-spacing: 0.3px; margin: 0; color: #FFFFFF; }
    .header-subtitle { font-size: 13.5px; color: #C9DCEE; margin-top: 4px; font-weight: 500; }

    /* CONTENEDORES PRINCIPALES */
    .st-key-panel_control, .st-key-sistema_control {
        background: #FFFFFF !important;
        border: 1px solid #DBE5EF;
        border-radius: 16px;
        padding: 18px 20px 20px;
        margin-bottom: 20px;
        box-shadow: 0 4px 14px rgba(15, 42, 70, 0.05);
    }
    .section-title {
        font-size: 1.02rem;
        font-weight: 800;
        color: #122F4C;
        margin-bottom: 14px;
        padding-bottom: 10px;
        border-bottom: 1px solid #E7EDF3;
    }

    /* FILA HORIZONTAL DE BLOQUES KPI */
    .kpi-row {
        display: grid;
        grid-template-columns: repeat(5, minmax(0, 1fr));
        align-items: stretch;
        gap: 12px;
    }
    @media (max-width: 1100px) { .kpi-row { grid-template-columns: repeat(3, minmax(0, 1fr)); } }
    @media (max-width: 700px) { .kpi-row { grid-template-columns: repeat(2, minmax(0, 1fr)); } }

    .kpi-block-card {
        background: #F4F8FC;
        border: 1px solid #E1E9F1;
        border-radius: 14px;
        padding: 12px 14px;
        display: flex;
        flex-direction: column;
        height: 100%;
        box-sizing: border-box;
    }
    .kpi-block-title {
        font-size: .68rem;
        font-weight: 800;
        text-transform: uppercase;
        letter-spacing: .5px;
        color: #5D7086;
        margin-bottom: 10px;
        white-space: nowrap;
    }
    .kpi-items { display: grid; gap: 8px; flex: 1; align-content: start; }
    .kpi-item {
        background: #FFFFFF;
        border: 1px solid #E6EDF4;
        border-left: 3.5px solid var(--tone);
        border-radius: 0 9px 9px 0;
        padding: 7px 11px;
    }
    .kpi-item-label {
        font-size: .63rem;
        font-weight: 750;
        text-transform: uppercase;
        color: #5D7086;
        letter-spacing: .2px;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
    }
    .kpi-item-value { font-size: 1.35rem; font-weight: 800; color: #102E4C; line-height: 1.15; }

    /* PESTAÑAS DISTRIBUIDAS PAREJAS */
    .st-key-sistema_control .stTabs [data-baseweb="tab-list"],
    .st-key-sistema_control .stTabs [role="tablist"] {
        display: flex !important;
        width: 100% !important;
        gap: 5px !important;
        flex-wrap: wrap;
        background-color: #EEF3F9;
        padding: 6px;
        border-radius: 10px;
    }
    .st-key-sistema_control .stTabs [data-baseweb="tab"],
    .st-key-sistema_control .stTabs button[role="tab"] {
        flex: 1 1 auto !important;
        min-width: 110px;
        justify-content: center !important;
        height: 38px;
        border-radius: 7px;
        font-size: 11.5px;
        font-weight: 650;
        color: #475569;
        padding: 0 8px !important;
        background: #FFFFFF;
        border: 1px solid #DFE7EF;
        text-align: center;
        white-space: nowrap;
    }
    .st-key-sistema_control .stTabs [aria-selected="true"] {
        background-color: #0E2A47 !important;
        color: #FFFFFF !important;
        border-color: #0E2A47 !important;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }

    div[data-testid="stDataFrame"], div[data-testid="stDataEditor"] {
        background-color: #FFFFFF !important;
        border-radius: 10px;
        border: 1px solid #E2E8F0;
        padding: 4px;
        margin-top: 10px !important;
    }
    div[data-testid="stExpander"] { background:#fff; border-color:#dbe5ef; border-radius:12px; }
    </style>
""",
    unsafe_allow_html=True,
)

# Constantes y Archivos
DB_FILE = "database_informes.json"
SOLICITUDES_FILE = "database_solicitudes.json"

COLUMNAS_EXCEL = [
    "ITEM POR MES", "IT2", "UNIDAD", "MES", "LINEAS", "CODIGO DE INFORME",
    "GRUPO DE TUBERÍAS", "SAP", "ALCANCE DEL SERVICIO", "NOTAS",
    "ESTADO - ELABORACIÓN ", "RESPONSABLE", "OBSERVACIÓN", "VALORIZACIÓN"
]

ORDEN_MESES = [
    "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
    "JULIO", "AGOSTO", "SETIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"
]

ESPECIALISTAS_LISTA = ["Jesús Rehkoff Díaz", "M. Paifa", "Julio Ponce", "Omar", "Christopher", "Timana", "Ingrid"]
REVISORES_PSAIM_LISTA = ["Franmary Gutierrez", "Alejandro Macury", "M. Paifa", "Julio Ponce", "Omar", "Christopher", "Timana", "Ingrid"]
PERSONAL_LISTA_BASE = ["M. Paifa", "Julio Ponce", "Omar", "Christopher", "Timana", "Ingrid", "Juan José", "Dante", "Jesús Rehkoff Díaz", "Franmary Gutierrez", "Alejandro Macury", "Otro Inspector"]

# Funciones de Limpieza y Normalización
def texto_normalizado(valor):
    if pd.isna(valor) or valor is None:
        return ""
    texto = str(valor).strip().upper()
    return texto.translate(str.maketrans("ÁÉÍÓÚÜÑ", "AEIOUUN"))

def texto_limpio(valor):
    if pd.isna(valor) or valor is None:
        return ""
    texto = str(valor).strip()
    return texto[:-2] if texto.endswith(".0") else texto

def separar_alcance_y_notas(alcance, notas=""):
    alcance_limpio = texto_limpio(alcance)
    notas_limpias = texto_limpio(notas)
    patron = re.compile(r"^\s*(LINEAS|VT\s*-\s*CIRCUITOS)\s*(?:[-–—:]\s*(.+))?$", flags=re.IGNORECASE)
    coincidencia = patron.match(alcance_limpio)
    if not coincidencia:
        return alcance_limpio, notas_limpias

    alcance_base = "LINEAS" if texto_normalizado(coincidencia.group(1)) == "LINEAS" else "VT-CIRCUITOS"
    nota_extraida = texto_limpio(coincidencia.group(2) or "")
    if nota_extraida and texto_normalizado(nota_extraida) not in texto_normalizado(notas_limpias):
        notas_limpias = f"{notas_limpias} | {nota_extraida}".strip(" |") if notas_limpias else nota_extraida
    return alcance_base, notas_limpias

def normalizar_base(df_entrada):
    df = df_entrada.copy()
    for columna in COLUMNAS_EXCEL:
        if columna not in df.columns:
            df[columna] = ""
    df = df[COLUMNAS_EXCEL].fillna("")

    for indice, fila in df.iterrows():
        estado = texto_limpio(fila["ESTADO - ELABORACIÓN "])
        responsable = texto_limpio(fila["RESPONSABLE"])
        if "-" in estado:
            estado_base, responsable_estado = estado.split("-", 1)
            df.at[indice, "ESTADO - ELABORACIÓN "] = estado_base.strip()
            if not responsable or texto_normalizado(responsable) in {"NAN", "NONE"}:
                df.at[indice, "RESPONSABLE"] = responsable_estado.strip()

        alcance, notas = separar_alcance_y_notas(fila["ALCANCE DEL SERVICIO"], fila.get("NOTAS", ""))
        df.at[indice, "ALCANCE DEL SERVICIO"] = alcance
        df.at[indice, "NOTAS"] = notas
        for columna in ["ITEM POR MES", "IT2", "SAP"]:
            df.at[indice, columna] = texto_limpio(fila[columna])
    return df

def es_codigo_provisional(codigo):
    codigo_normalizado = texto_normalizado(codigo)
    return codigo_normalizado in {"", "-"} or any(
        texto in codigo_normalizado for texto in ["PENDIENTE ASIGNAR", "PENDIENTE DE ASIGNAR", "POR ASIGNAR"]
    )

def es_correccion_psaim(observacion):
    observacion = texto_normalizado(observacion)
    return "PSAIM" in observacion and any(
        texto in observacion for texto in ["CORRECCION", "CORREGIR", "CORREGIDO", "CORREGIDA"]
    )

def es_revision_fiabilidad(observacion):
    obs = texto_normalizado(observacion)
    return "FIABILIDAD" in obs and "REVISION" in obs

def es_pendiente_inspeccion(fila):
    estado = texto_normalizado(fila.get("ESTADO - ELABORACIÓN ", ""))
    notas = texto_normalizado(fila.get("NOTAS", ""))
    observacion = texto_normalizado(fila.get("OBSERVACIÓN", ""))
    return any(
        texto in estado or texto in notas or texto in observacion
        for texto in ["PENDIENTE COMPLETAR INSPECCION", "PENDIENTE INSPECCION", "FALTA CARPETA", "COMPLETAR INSPECCION"]
    )

@st.cache_data(ttl=5, show_spinner=False)
def cargar_datos():
    descargar_archivo_de_drive(DB_FILE, DB_FILE)
    if not os.path.exists(DB_FILE):
        return pd.DataFrame(columns=COLUMNAS_EXCEL)
    try:
        with open(DB_FILE, "r", encoding="utf-8") as archivo:
            return normalizar_base(pd.DataFrame(json.load(archivo)))
    except Exception:
        return pd.DataFrame(columns=COLUMNAS_EXCEL)

def guardar_datos(df):
    normalizar_base(df).to_json(DB_FILE, orient="records", force_ascii=False)
    subir_a_drive_en_segundo_plano(DB_FILE, DB_FILE)
    st.cache_data.clear()

@st.cache_data(ttl=5, show_spinner=False)
def cargar_solicitudes():
    descargar_archivo_de_drive(SOLICITUDES_FILE, SOLICITUDES_FILE)
    if not os.path.exists(SOLICITUDES_FILE):
        return []
    try:
        with open(SOLICITUDES_FILE, "r", encoding="utf-8") as archivo:
            return json.load(archivo)
    except Exception:
        return []

def guardar_solicitudes(solicitudes):
    with open(SOLICITUDES_FILE, "w", encoding="utf-8") as archivo:
        json.dump(solicitudes, archivo, ensure_ascii=False)
    subir_a_drive_en_segundo_plano(SOLICITUDES_FILE, SOLICITUDES_FILE)
    st.cache_data.clear()

def registrar_solicitud(tipo, codigo, grupo, solicitante):
    solicitudes = cargar_solicitudes()
    repetida = any(
        solicitud["codigo"] == codigo and solicitud["grupo"] == grupo and solicitud["tipo"] == tipo and solicitud["estado"] == "PENDIENTE"
        for solicitud in solicitudes
    )
    if repetida:
        return False, "Ya existe una solicitud pendiente para este informe."
    siguiente_id = max((solicitud.get("id", 0) for solicitud in solicitudes), default=0) + 1
    solicitudes.append({
        "id": siguiente_id, "tipo": tipo, "codigo": codigo, "grupo": grupo,
        "solicitante": solicitante, "estado": "PENDIENTE"
    })
    guardar_solicitudes(solicitudes)
    return True, "Solicitud enviada al administrador."

def excel_con_formato(df, nombre_hoja="CONTROL"):
    salida = io.BytesIO()
    datos = df.copy().fillna("")
    with pd.ExcelWriter(salida, engine="openpyxl") as escritor:
        datos.to_excel(escritor, index=False, sheet_name=nombre_hoja[:31])
        hoja = escritor.book[nombre_hoja[:31]]
        hoja.freeze_panes = "A2"
        hoja.row_dimensions[1].height = 30
        relleno = PatternFill("solid", fgColor="0E2A47")
        for celda in hoja[1]:
            celda.font = Font(color="FFFFFF", bold=True)
            celda.fill = relleno
            celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for columna in hoja.columns:
            letra = get_column_letter(columna[0].column)
            ancho = max(len(texto_limpio(celda.value)) for celda in columna)
            hoja.column_dimensions[letra].width = min(max(ancho + 2, 12), 42)
            for celda in columna[1:]:
                celda.alignment = Alignment(vertical="top", wrap_text=True)

        ultima_columna = get_column_letter(max(1, len(datos.columns)))
        ultima_fila = len(datos) + 1
        if len(datos.columns) and len(datos):
            tabla = Table(displayName=f"Tabla_{datetime.now():%H%M%S%f}", ref=f"A1:{ultima_columna}{ultima_fila}")
            tabla.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            hoja.add_table(tabla)
        else:
            hoja.auto_filter.ref = f"A1:{ultima_columna}1"
    salida.seek(0)
    return salida.getvalue()

def boton_descarga_excel(df, archivo, etiqueta="Descargar Excel"):
    st.download_button(
        etiqueta,
        data=excel_con_formato(df),
        file_name=archivo,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        icon=":material/download:",
        use_container_width=False,
    )

def senal_visual(fila):
    notas = texto_normalizado(fila.get("NOTAS", ""))
    observacion = texto_normalizado(fila.get("OBSERVACIÓN", ""))
    val = texto_normalizado(fila.get("VALORIZACIÓN", ""))
    if "RETIRADO" in notas or "RETIRADO" in observacion or val == "RETIRADO":
        return "🔴 Retirado"
    if val == "SI":
        return "🟢 Valorizado (SI)"
    if "FALTA CARPETA" in notas or "PENDIENTE INSPECCION" in notas:
        return "🟡 Pend. inspección"
    if "INSPECCION COMPLEMENTARIA" in notas:
        return "🔵 Inspección complem."
    return "⚪ Sin alerta"

# Carga Inicial de Datos
if "df_data" not in st.session_state:
    st.session_state.df_data = cargar_datos()

df = normalizar_base(st.session_state.df_data)

# Banner Principal
st.html("""
    <div class="header-banner">
        <div class="header-title">CONTROL INTERNO DE INFORMES - ADEMINSAC</div>
        <div class="header-subtitle">Sistema de Monitoreo de Inspección Técnicas y Valorización | Refinería La Pampilla</div>
    </div>
""")

# Expander de Carga / Respaldo
with st.expander("⚙️ Gestión de datos: cargar, restaurar y descargar respaldo", expanded=False):
    carga, respaldo = st.columns([1.15, 0.85], vertical_alignment="bottom")
    with carga:
        st.subheader("Cargar base de datos")
        archivo_excel = st.file_uploader("Selecciona un archivo Excel", type=["xlsx", "xlsm"])
        if archivo_excel and st.button("Reemplazar base de datos", icon=":material/upload:", type="primary"):
            try:
                libro = pd.ExcelFile(archivo_excel)
                hoja = "CONTROL" if "CONTROL" in libro.sheet_names else libro.sheet_names[0]
                df_cargado = pd.read_excel(libro, sheet_name=hoja)
                mapa = {str(columna).strip().upper(): columna for columna in df_cargado.columns}
                renombre = {mapa[columna.strip().upper()]: columna for columna in COLUMNAS_EXCEL if columna.strip().upper() in mapa}
                df_cargado = df_cargado.rename(columns=renombre)
                st.session_state.df_data = normalizar_base(df_cargado)
                guardar_datos(st.session_state.df_data)
                st.success("Base de datos cargada, migrada y guardada en Google Drive.")
                st.rerun()
            except Exception as error:
                st.error(f"No se pudo cargar el Excel: {error}")
    with respaldo:
        st.subheader("Descargar respaldo actual")
        if not df.empty:
            boton_descarga_excel(df, "Respaldo_Control_Informes.xlsx", "Descargar copia en Excel")
        else:
            st.caption("Carga una base de datos para generar el respaldo.")

if df.empty:
    st.info("Carga un archivo Excel desde Gestión de datos para iniciar el control.", icon=":material/info:")
    st.stop()

# OPTIMIZACIÓN CON CACHÉ DE PROCESAMIENTO DE DATOS
@st.cache_data(show_spinner=False)
def procesar_agrupaciones_y_kpis(df_input):
    mascara_retirado = df_input["OBSERVACIÓN"].apply(lambda v: "RETIRADO" in texto_normalizado(v)) | \
                       df_input["NOTAS"].apply(lambda v: "RETIRADO" in texto_normalizado(v)) | \
                       df_input["VALORIZACIÓN"].apply(lambda v: texto_normalizado(v) == "RETIRADO")
    df_activos = df_input[~mascara_retirado].copy()

    df_activos["CLAVE_GLOBAL"] = df_activos.apply(
        lambda fila: f"{texto_limpio(fila['MES'])}|SIN-CODIGO-GRUPO|{texto_normalizado(fila['GRUPO DE TUBERÍAS'])}"
        if es_codigo_provisional(fila["CODIGO DE INFORME"])
        else f"{texto_limpio(fila['MES'])}|{texto_limpio(fila['CODIGO DE INFORME'])}",
        axis=1,
    )

    mask_psaim = df_activos["OBSERVACIÓN"].apply(es_correccion_psaim)
    mask_pend_inspeccion = df_activos.apply(es_pendiente_inspeccion, axis=1)
    mask_pend_elaboracion = df_activos["ESTADO - ELABORACIÓN "].apply(texto_normalizado).str.contains("PENDIENTE ELABORACION")

    df_psaim = df_activos[mask_psaim]
    df_pend_inspeccion = df_activos[mask_pend_inspeccion]
    df_pend_asignacion = df_activos[mask_pend_elaboracion]
    df_en_proceso = df_activos[df_activos["ESTADO - ELABORACIÓN "].apply(texto_normalizado).str.contains("EN PROCESO") & ~mask_pend_inspeccion]

    unicos, psaim_unicos = set(), set()
    unicos_finalizados = set()

    por_mes = {"valorizados": {}, "pendientes": {}, "ademinsac": {}, "fiabilidad": {}, "psaim": {}}
    detalle_pendientes = {}
    revision_fiabilidad = revision_especialista_pendiente = revision_especialista = 0

    for _, fila in df_activos.iterrows():
        mes = texto_limpio(fila["MES"])
        codigo = texto_limpio(fila["CODIGO DE INFORME"])
        grupo = texto_limpio(fila["GRUPO DE TUBERÍAS"])
        observacion = texto_limpio(fila["OBSERVACIÓN"])
        estado_elab = texto_normalizado(fila["ESTADO - ELABORACIÓN "])
        clave = fila["CLAVE_GLOBAL"]
        if not mes or not grupo:
            continue
        if not es_codigo_provisional(codigo) and es_correccion_psaim(observacion):
            clave_psaim = f"{mes}|{codigo}"
            if clave_psaim not in psaim_unicos:
                psaim_unicos.add(clave_psaim)
                por_mes["psaim"][mes] = por_mes["psaim"].get(mes, 0) + 1
        if clave in unicos:
            continue
        unicos.add(clave)
        
        if "FINALIZADO" in estado_elab or "100%" in estado_elab:
            unicos_finalizados.add(clave)

        for clave_mes in ["valorizados", "pendientes", "ademinsac", "fiabilidad"]:
            por_mes[clave_mes].setdefault(mes, 0)
        if texto_normalizado(fila["VALORIZACIÓN"]) == "SI":
            por_mes["valorizados"][mes] += 1
            continue
        por_mes["pendientes"][mes] += 1
        observacion_norm = texto_normalizado(observacion)
        if es_revision_fiabilidad(observacion):
            revision_fiabilidad += 1
        if "PENDIENTE REVISION POR EL ESPECIALISTA" in observacion_norm:
            revision_especialista_pendiente += 1
        if ("REV. POR EL ESPECIALISTA" in observacion_norm or "REVISION POR EL ESPECIALISTA" in observacion_norm) and "PENDIENTE" not in observacion_norm:
            revision_especialista += 1
        if "ADEMINSAC" in observacion_norm:
            por_mes["ademinsac"][mes] += 1
        else:
            por_mes["fiabilidad"][mes] += 1
        etiqueta = "(En blanco)" if not observacion else observacion
        detalle_pendientes[(mes, etiqueta)] = detalle_pendientes.get((mes, etiqueta), 0) + 1

    total_inf_unicos = len(unicos)
    tot_finalizados = len(unicos_finalizados)
    tot_pendientes_elaborar = max(0, total_inf_unicos - tot_finalizados)

    tot_valorizados = sum(por_mes["valorizados"].values())
    val_para_asignar = df_pend_asignacion["CLAVE_GLOBAL"].nunique()
    val_en_proceso = df_en_proceso["CLAVE_GLOBAL"].nunique()
    val_pend_inspeccion = df_pend_inspeccion["CLAVE_GLOBAL"].nunique()
    val_psaim = sum(por_mes["psaim"].values())

    kpis = {
        "total_inf_unicos": total_inf_unicos,
        "tot_finalizados": tot_finalizados,
        "tot_pendientes_elaborar": tot_pendientes_elaborar,
        "tot_valorizados": tot_valorizados,
        "val_para_asignar": val_para_asignar,
        "val_en_proceso": val_en_proceso,
        "val_pend_inspeccion": val_pend_inspeccion,
        "val_psaim": val_psaim,
        "revision_especialista": revision_especialista,
        "revision_especialista_pendiente": revision_especialista_pendiente,
        "revision_fiabilidad": revision_fiabilidad
    }

    return mascara_retirado, df_activos, df_psaim, df_pend_inspeccion, df_pend_asignacion, df_en_proceso, kpis, detalle_pendientes

mascara_retirado, df_activos, df_psaim, df_pend_inspeccion, df_pend_asignacion, df_en_proceso, kpis, detalle_pendientes = procesar_agrupaciones_y_kpis(df)

# RENDERIZADO DEL PANEL DE CONTROL
def item_kpi(titulo, valor, color):
    return (
        f"<div class='kpi-item' style='--tone:{color}'>"
        f"<div class='kpi-item-label'>{titulo}</div>"
        f"<div class='kpi-item-value'>{valor}</div>"
        f"</div>"
    )

def bloque_kpi(titulo_bloque, emoji, items):
    filas = "".join(item_kpi(*i) for i in items)
    return (
        f"<div class='kpi-block-card'>"
        f"<div class='kpi-block-title'>{emoji} {titulo_bloque}</div>"
        f"<div class='kpi-items'>{filas}</div></div>"
    )

panel_control = st.container(key="panel_control")
panel_control.markdown("<div class='section-title'>📊 Panel de control de informes</div>", unsafe_allow_html=True)

bloques_html = "".join([
    bloque_kpi("Bloque general", "📊", [
        ("Informes totales", kpis["total_inf_unicos"], "#173F67"),
        ("Informes finalizados", kpis["tot_finalizados"], "#159A68"),
        ("Pendientes elaborar", kpis["tot_pendientes_elaborar"], "#E38921"),
    ]),
    bloque_kpi("Bloque gabinete", "📁", [
        ("En proceso", kpis["val_en_proceso"], "#7B61C9"),
        ("Pend. asignar", kpis["val_para_asignar"], "#D54D9D"),
        ("Correc. PSAIM", kpis["val_psaim"], "#C89716"),
    ]),
    bloque_kpi("Bloque especialista", "👤", [
        ("Revisados", kpis["revision_especialista"], "#168EAE"),
        ("Por revisar", kpis["revision_especialista_pendiente"], "#5564D8"),
    ]),
    bloque_kpi("Bloque campo", "📝", [
        ("Pend. inspección", kpis["val_pend_inspeccion"], "#D8534F"),
    ]),
    bloque_kpi("Bloque cliente", "🏢", [
        ("Valorizados", kpis["tot_valorizados"], "#159A68"),
        ("En revisión", kpis["revision_fiabilidad"], "#159D99"),
    ]),
])

panel_control.markdown(f"<div class='kpi-row'>{bloques_html}</div>", unsafe_allow_html=True)

# SISTEMA DE CONTROL Y RESÚMENES
solicitudes_activas = [solicitud for solicitud in cargar_solicitudes() if solicitud["estado"] == "PENDIENTE"]

sistema_control = st.container(key="sistema_control")
sistema_control.markdown("<div class='section-title'>🗂️ Sistema de control y resúmenes</div>", unsafe_allow_html=True)

tabs = sistema_control.tabs([
    f"🔔 Admin ({len(solicitudes_activas)})",
    "📋 Tabla general",
    "📇 Pend. asignar",
    "🔄 En proceso",
    "⏳ Pend. inspección",
    "🔍 Rev. fiabilidad",
    "👤 Revisión especialista",
    "🛠️ Correc. PSAIM",
    "📊 Resumen por mes",
])

# 1. ADMIN
with tabs[0]:
    @st.fragment
    def vista_admin():
        st.subheader("Bandeja de aprobación")
        sols = [s for s in cargar_solicitudes() if s["estado"] == "PENDIENTE"]
        if not sols:
            st.success("No hay solicitudes pendientes.", icon=":material/check_circle:")
        for solicitud in sols:
            with st.container(border=True):
                informacion, aprobar, rechazar = st.columns([5, 1, 1], vertical_alignment="center")
                informacion.write(f"**{solicitud['tipo']}**  \nCódigo: `{solicitud['codigo']}` | Grupo: `{solicitud['grupo']}` | Solicitante: {solicitud['solicitante']}")
                if aprobar.button("Aprobar", key=f"aprobar_{solicitud['id']}", icon=":material/check:"):
                    mascara_base = (df["CODIGO DE INFORME"] == solicitud["codigo"]) & (df["GRUPO DE TUBERÍAS"] == solicitud["grupo"])
                    
                    if solicitud["tipo"] == "INFORME COMPLETADO (GABINETE)":
                        mascara_no_retirado = mascara_base & ~mascara_retirado
                        df.loc[mascara_no_retirado, "ESTADO - ELABORACIÓN "] = "Finalizado"
                        df.loc[mascara_no_retirado, "OBSERVACIÓN"] = "Pendiente revisión por el especialista - Ademinsac"
                    elif solicitud["tipo"] == "CORRECCIÓN PSAIM":
                        df.loc[mascara_base, "OBSERVACIÓN"] = "PSAIM CORREGIDO"
                        df.loc[mascara_base, "ESTADO - ELABORACIÓN "] = "En proceso"
                    elif solicitud["tipo"] == "REVISIÓN ESPECIALISTA":
                        df.loc[mascara_base, "OBSERVACIÓN"] = "INFORME REVISADO POR ESPECIALISTA"
                    
                    solicitudes = cargar_solicitudes()
                    for item in solicitudes:
                        if item["id"] == solicitud["id"]:
                            item["estado"] = "APROBADO"
                    guardar_solicitudes(solicitudes)
                    st.session_state.df_data = normalizar_base(df)
                    guardar_datos(st.session_state.df_data)
                    st.rerun()
                if rechazar.button("Rechazar", key=f"rechazar_{solicitud['id']}", icon=":material/close:"):
                    solicitudes = cargar_solicitudes()
                    for item in solicitudes:
                        if item["id"] == solicitud["id"]:
                            item["estado"] = "RECHAZADO"
                    guardar_solicitudes(solicitudes)
                    st.rerun()
    vista_admin()

# 2. TABLA GENERAL
with tabs[1]:
    @st.fragment
    def vista_tabla_general():
        filtros = st.columns([1, 1, 2])
        meses = ["Todos"] + sorted(
            {texto_limpio(m).upper() for m in df["MES"] if texto_limpio(m)},
            key=lambda m: ORDEN_MESES.index(m) if m in ORDEN_MESES else 99,
        )
        mes = filtros[0].selectbox("Filtrar mes", meses)
        alcance = filtros[1].selectbox("Alcance del servicio", ["Todos", "LINEAS", "VT-CIRCUITOS"])
        consulta = filtros[2].text_input("Buscar por líneas, código, grupo, SAP o notas", icon=":material/search:")
        
        df_vista = df.copy()
        if mes != "Todos":
            df_vista = df_vista[df_vista["MES"].apply(lambda v: texto_normalizado(v) == mes)]
        if alcance != "Todos":
            df_vista = df_vista[df_vista["ALCANCE DEL SERVICIO"].apply(texto_normalizado) == alcance]
        if consulta.strip():
            consulta_norm = texto_normalizado(consulta)
            columnas_busqueda = ["LINEAS", "CODIGO DE INFORME", "GRUPO DE TUBERÍAS", "SAP", "NOTAS"]
            mascara_busqueda = df_vista[columnas_busqueda].apply(
                lambda fila: any(consulta_norm in texto_normalizado(v) for v in fila), axis=1
            )
            df_vista = df_vista[mascara_busqueda]

        df_vista = df_vista.map(texto_limpio)
        df_vista["VALORIZACIÓN"] = df_vista["VALORIZACIÓN"].apply(
            lambda v: "SI" if texto_normalizado(v) == "SI" else ("Retirado" if texto_normalizado(v) == "RETIRADO" else "Pendiente")
        )
        df_vista.insert(0, "SEÑAL", df_vista.apply(senal_visual, axis=1))
        
        encabezados = {
            "SEÑAL": st.column_config.TextColumn("Señal", width=190, disabled=True, pinned=True),
            "ITEM POR MES": st.column_config.TextColumn("Item", width=70),
            "IT2": st.column_config.TextColumn("IT2", width=55),
            "UNIDAD": st.column_config.TextColumn("Unidad", width=65),
            "MES": st.column_config.TextColumn("Mes", width=80),
            "LINEAS": st.column_config.TextColumn("Líneas", width=180),
            "CODIGO DE INFORME": st.column_config.TextColumn("Código de informe", width=190),
            "GRUPO DE TUBERÍAS": st.column_config.TextColumn("Grupo de tuberías", width=180),
            "SAP": st.column_config.TextColumn("SAP", width=85),
            "ALCANCE DEL SERVICIO": st.column_config.TextColumn("Alcance", width=120),
            "NOTAS": st.column_config.TextColumn("Notas", width=170),
            "ESTADO - ELABORACIÓN ": st.column_config.TextColumn("Estado de elaboración", width=190),
            "RESPONSABLE": st.column_config.TextColumn("Responsable", width=135),
            "OBSERVACIÓN": st.column_config.TextColumn("Observación", width=280),
            "VALORIZACIÓN": st.column_config.SelectboxColumn(
                "Valorización", options=["Pendiente", "SI", "Retirado"], required=True, width=145
            ),
        }
        
        st.html("""
            <div style="background-color: #f1f5f9; border: 1px solid #cbd5e1; padding: 8px 14px; border-radius: 8px; font-size: 0.85rem; font-weight: 600; color: #102E4C; margin-bottom: 10px; display: inline-block;">
                🟢 Valorizado (SI) &nbsp;&nbsp;|&nbsp;&nbsp; 🟡 Pendiente de inspección o falta carpeta &nbsp;&nbsp;|&nbsp;&nbsp; 🔵 Inspección complementaria &nbsp;&nbsp;|&nbsp;&nbsp; 🔴 Retirado
            </div>
        """)
        
        with st.expander("⚡ Valorización masiva por Código de Informe", expanded=False):
            col_cod, col_est, col_btn = st.columns([3, 2, 1], vertical_alignment="bottom")
            
            codigos_disponibles = sorted([
                c for c in df["CODIGO DE INFORME"].unique() 
                if c and str(c) != "-" and not es_codigo_provisional(c)
            ])
            
            codigo_sel = col_cod.selectbox("Seleccionar Código de Informe", codigos_disponibles, key="val_masiva_cod")
            estado_sel = col_est.selectbox("Estado a aplicar", ["SI", "Pendiente"], key="val_masiva_est")
            
            if col_btn.button("Aplicar a todo", icon=":material/done_all:", type="primary"):
                mascara_objetivo = (df["CODIGO DE INFORME"] == codigo_sel) & ~mascara_retirado
                
                df.loc[mascara_objetivo, "VALORIZACIÓN"] = estado_sel
                if estado_sel == "SI":
                    df.loc[mascara_objetivo, "OBSERVACIÓN"] = ""
                
                st.session_state.df_data = normalizar_base(df)
                guardar_datos(st.session_state.df_data)
                st.toast(f"Valorización actualizada a '{estado_sel}' para {codigo_sel}", icon="✅")
                st.rerun()

        boton_descarga_excel(df_vista, "Tabla_general_informes.xlsx", "Descargar tabla general")

        editado = st.data_editor(
            df_vista,
            column_config=encabezados,
            hide_index=True,
            width="stretch",
            height=600,
            disabled=["SEÑAL"],
            key="editor_tabla_general",
        )
        
        if st.button("Guardar cambios", key="guardar_tabla", icon=":material/save:", type="primary"):
            df_actualizado = editado.drop(columns=["SEÑAL"], errors="ignore")
            
            mascara_si = df_actualizado["VALORIZACIÓN"].apply(lambda x: texto_normalizado(x) == "SI")
            df_actualizado.loc[mascara_si, "OBSERVACIÓN"] = ""
            
            st.session_state.df_data.update(df_actualizado)
            guardar_datos(st.session_state.df_data)
            
            st.toast("¡Cambios guardados con éxito!", icon="💾")
            st.rerun()

    vista_tabla_general()

# AUXILIARES PARA AGRUPACIÓN DE TABLAS
def tabla_agrupada(df_origen, columnas, nombre_archivo, nombre_hoja):
    if df_origen.empty:
        st.info("No hay registros para mostrar.", icon=":material/info:")
        return pd.DataFrame()
    tabla = df_origen.groupby(columnas, as_index=False, dropna=False).agg(LINEAS=("LINEAS", "count")).fillna("")
    tabla.index = range(1, len(tabla) + 1)
    boton_descarga_excel(tabla, nombre_archivo, "Descargar Excel")
    st.dataframe(tabla, width="stretch", hide_index=False, height=600)
    return tabla

def mostrar_resumen(df_resumen, nombre_archivo, es_metricas=False):
    if df_resumen.empty:
        st.info("No hay registros para mostrar.", icon=":material/info:")
        return
    
    df_mostrar = df_resumen.copy()
    if es_metricas:
        tot_informes = df_mostrar["TOTAL INFORMES"].sum()
        tot_elaborados = df_mostrar["INFORMES ELABORADOS"].sum()
        tot_pendientes = df_mostrar["PENDIENTES POR ELABORAR"].sum()
        pct_total = (tot_elaborados / tot_informes * 100) if tot_informes > 0 else 0.0

        fila_total = pd.DataFrame({
            "MES": ["TOTAL"],
            "TOTAL INFORMES": [tot_informes],
            "INFORMES ELABORADOS": [tot_elaborados],
            "PENDIENTES POR ELABORAR": [tot_pendientes],
            "% AVANCE ELABORACIÓN": [f"{pct_total:.1f}%"]
        })
        df_mostrar["% AVANCE ELABORACIÓN"] = df_mostrar["% AVANCE ELABORACIÓN"].apply(
            lambda v: f"{v:.1f}%" if isinstance(v, (int, float)) else str(v)
        )
        df_mostrar = pd.concat([df_mostrar, fila_total], ignore_index=True)
    elif "CANTIDAD" in df_mostrar.columns:
        tot_cantidad = df_mostrar["CANTIDAD"].sum()
        fila_total = pd.DataFrame({
            "MES": ["TOTAL"],
            "OBSERVACIÓN PENDIENTE": ["-"],
            "CANTIDAD": [tot_cantidad]
        })
        df_mostrar = pd.concat([df_mostrar, fila_total], ignore_index=True)

    df_mostrar.index = range(1, len(df_mostrar) + 1)
    boton_descarga_excel(df_mostrar, nombre_archivo, "Descargar Excel")
    st.dataframe(df_mostrar, width="stretch", hide_index=False, height=600)

# 3. PENDIENTE ASIGNAR
with tabs[2]:
    tabla_agrupada(
        df_pend_asignacion,
        ["MES", "ESTADO - ELABORACIÓN ", "RESPONSABLE", "GRUPO DE TUBERÍAS", "CODIGO DE INFORME"],
        "Pendientes_asignar.xlsx", "PEND_ASIGNAR"
    )

# 4. EN PROCESO
with tabs[3]:
    @st.fragment
    def vista_en_proceso():
        tabla_proceso = tabla_agrupada(
            df_en_proceso,
            ["MES", "ESTADO - ELABORACIÓN ", "RESPONSABLE", "GRUPO DE TUBERÍAS", "CODIGO DE INFORME"],
            "En_proceso.xlsx", "EN_PROCESO"
        )
        if not tabla_proceso.empty:
            responsables = sorted(set(PERSONAL_LISTA_BASE + [texto_limpio(valor) for valor in df["RESPONSABLE"] if texto_limpio(valor)]))
            codigo, inspector, enviar = st.columns([2, 2, 1], vertical_alignment="bottom")
            codigo_seleccionado = codigo.selectbox("Código", tabla_proceso["CODIGO DE INFORME"].unique(), key="proceso_codigo")
            inspector_seleccionado = inspector.selectbox("Inspector", responsables, key="proceso_inspector")
            if enviar.button("Enviar al 100%", icon=":material/send:"):
                grupo = tabla_proceso.loc[tabla_proceso["CODIGO DE INFORME"] == codigo_seleccionado, "GRUPO DE TUBERÍAS"].iloc[0]
                correcto, mensaje = registrar_solicitud("INFORME COMPLETADO (GABINETE)", codigo_seleccionado, grupo, inspector_seleccionado)
                if correcto:
                    st.success(mensaje)
                    st.rerun()
                else:
                    st.warning(mensaje)
    vista_en_proceso()

# 5. PENDIENTE INSPECCIÓN
with tabs[4]:
    tabla_agrupada(
        df_pend_inspeccion,
        ["MES", "ESTADO - ELABORACIÓN ", "RESPONSABLE", "GRUPO DE TUBERÍAS", "CODIGO DE INFORME"],
        "Pendientes_inspeccion.xlsx", "PEND_INSPECCION"
    )

# 6. REV. FIABILIDAD
with tabs[5]:
    df_fiabilidad = df_activos[df_activos["OBSERVACIÓN"].apply(es_revision_fiabilidad)]
    tabla_agrupada(
        df_fiabilidad,
        ["MES", "ESTADO - ELABORACIÓN ", "RESPONSABLE", "GRUPO DE TUBERÍAS", "CODIGO DE INFORME", "OBSERVACIÓN"],
        "Revision_fiabilidad.xlsx", "REV_FIABILIDAD"
    )

# 7. REVISIÓN ESPECIALISTA
def vista_revision_especialista(condicion, archivo, llave):
    df_revision = df_activos[df_activos["OBSERVACIÓN"].apply(condicion)]
    tabla_revision = tabla_agrupada(
        df_revision,
        ["MES", "ESTADO - ELABORACIÓN ", "RESPONSABLE", "GRUPO DE TUBERÍAS", "CODIGO DE INFORME", "OBSERVACIÓN"],
        archivo, llave
    )
    if tabla_revision.empty:
        return
    codigo, especialista, enviar = st.columns([2, 2, 1], vertical_alignment="bottom")
    codigo_seleccionado = codigo.selectbox("Código", tabla_revision["CODIGO DE INFORME"].unique(), key=f"codigo_{llave}")
    especialista_seleccionado = especialista.selectbox("Especialista", ESPECIALISTAS_LISTA, key=f"especialista_{llave}")
    if enviar.button("Enviar a revisión", key=f"enviar_{llave}", icon=":material/send:"):
        grupo = tabla_revision.loc[tabla_revision["CODIGO DE INFORME"] == codigo_seleccionado, "GRUPO DE TUBERÍAS"].iloc[0]
        correcto, mensaje = registrar_solicitud("REVISIÓN ESPECIALISTA", codigo_seleccionado, grupo, especialista_seleccionado)
        if correcto:
            st.success(mensaje)
            st.rerun()
        else:
            st.warning(mensaje)

with tabs[6]:
    st.subheader("Revisión especialista")
    @st.fragment
    def vista_sub_especialista():
        opcion_especialista = st.radio(
            "Seleccionar tipo de vista:",
            ["Pendientes de revisión", "Revisados por el especialista"],
            horizontal=True,
            key="radio_especialistas"
        )
        if opcion_especialista == "Pendientes de revisión":
            vista_revision_especialista(
                lambda valor: "PENDIENTE REVISION POR EL ESPECIALISTA" in texto_normalizado(valor),
                "Pendientes_revision_especialista.xlsx", "PEND_REV_ESP"
            )
        else:
            vista_revision_especialista(
                lambda valor: ("REV. POR EL ESPECIALISTA" in texto_normalizado(valor) or "REVISION POR EL ESPECIALISTA" in texto_normalizado(valor))
                and "PENDIENTE" not in texto_normalizado(valor),
                "Revision_por_especialista.xlsx", "REV_POR_ESP"
            )
    vista_sub_especialista()

# 8. CORRECCIÓN PSAIM
with tabs[7]:
    @st.fragment
    def vista_correc_psaim():
        df_psaim_lineas = df_psaim[df_psaim["ALCANCE DEL SERVICIO"].apply(texto_normalizado) == "LINEAS"].copy()
        columnas_psaim = [
            "MES", "ESTADO - ELABORACIÓN ", "RESPONSABLE", "ITEM POR MES", "IT2",
            "LINEAS", "GRUPO DE TUBERÍAS", "CODIGO DE INFORME", "NOTAS", "OBSERVACIÓN"
        ]
        tabla_psaim = tabla_agrupada(
            df_psaim_lineas,
            columnas_psaim[:-3] + ["CODIGO DE INFORME", "NOTAS", "OBSERVACIÓN"],
            "Correccion_PSAIM.xlsx", "CORRECCION_PSAIM"
        )
        if not tabla_psaim.empty:
            codigo, revisor, enviar = st.columns([2, 2, 1], vertical_alignment="bottom")
            codigo_seleccionado = codigo.selectbox("Código", tabla_psaim["CODIGO DE INFORME"].unique(), key="psaim_codigo")
            revisor_seleccionado = revisor.selectbox("Revisor PSAIM", REVISORES_PSAIM_LISTA, key="psaim_revisor")
            if enviar.button("PSAIM corregido", icon=":material/check_circle:"):
                grupo = tabla_psaim.loc[tabla_psaim["CODIGO DE INFORME"] == codigo_seleccionado, "GRUPO DE TUBERÍAS"].iloc[0]
                correcto, mensaje = registrar_solicitud("CORRECCIÓN PSAIM", codigo_seleccionado, grupo, revisor_seleccionado)
                if correcto:
                    st.success(mensaje)
                    st.rerun()
                else:
                    st.warning(mensaje)
    vista_correc_psaim()

# 9. RESUMEN POR MES
@st.cache_data(show_spinner=False)
def generar_resumenes_mes(df_activos_input, detalle_pendientes_input):
    filas_elaboracion = []
    meses_unicos = sorted(
        set(df_activos_input["MES"].apply(texto_limpio)),
        key=lambda m: ORDEN_MESES.index(m.upper()) if m.upper() in ORDEN_MESES else 99,
    )

    for mes in meses_unicos:
        if not mes:
            continue
        df_mes = df_activos_input[df_activos_input["MES"].apply(lambda v: texto_limpio(v) == mes)]
        df_mes_unicos = df_mes.drop_duplicates(subset=["CLAVE_GLOBAL"])
        total_informes = len(df_mes_unicos)
        
        elaborados = len(
            df_mes_unicos[
                df_mes_unicos["ESTADO - ELABORACIÓN "].apply(
                    lambda v: "FINALIZADO" in texto_normalizado(v) or "100%" in texto_normalizado(v)
                )
            ]
        )
        pendientes_elaborar = total_informes - elaborados
        porcentaje = round((elaborados / total_informes * 100), 1) if total_informes > 0 else 0.0
        
        filas_elaboracion.append({
            "MES": mes,
            "TOTAL INFORMES": total_informes,
            "INFORMES ELABORADOS": elaborados,
            "PENDIENTES POR ELABORAR": pendientes_elaborar,
            "% AVANCE ELABORACIÓN": porcentaje,
        })

    df_metricas_elaboracion = pd.DataFrame(filas_elaboracion)

    df_t4 = pd.DataFrame([
        {"MES": mes, "OBSERVACIÓN PENDIENTE": observacion, "CANTIDAD": cantidad}
        for (mes, observacion), cantidad in detalle_pendientes_input.items()
    ])
    if not df_t4.empty:
        df_t4["ORDEN"] = df_t4["MES"].apply(
            lambda valor: ORDEN_MESES.index(texto_normalizado(valor)) if texto_normalizado(valor) in ORDEN_MESES else 99
        )
        df_t4 = df_t4.sort_values(["ORDEN", "CANTIDAD"], ascending=[True, False]).drop(columns="ORDEN")

    return df_metricas_elaboracion, df_t4

df_metricas_elaboracion, df_t4 = generar_resumenes_mes(df_activos, detalle_pendientes)

with tabs[8]:
    st.subheader("Resumen por mes")
    @st.fragment
    def vista_sub_resumen():
        opcion_resumen = st.radio(
            "Seleccionar tipo de vista:",
            ["Métricas por mes", "Detalle pendientes por mes / observación"],
            horizontal=True,
            key="radio_resumen"
        )
        if opcion_resumen == "Métricas por mes":
            mostrar_resumen(df_metricas_elaboracion, "Metricas_Elaboracion_Por_Mes.xlsx", es_metricas=True)
        else:
            mostrar_resumen(df_t4, "Pendientes_mes_observacion_T4.xlsx", es_metricas=False)

    vista_sub_resumen()
